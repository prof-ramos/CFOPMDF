from __future__ import annotations

import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

INPUT_MD = Path("/Users/gabrielramos/CFOPMDF/CONTEUDO_PROGRAMATICO/CONTEUDO_PMDF.md")
OUTPUT_XLSX = Path("/Users/gabrielramos/CFOPMDF/CONTEUDO_PROGRAMATICO/CONTEUDO_PMDF.xlsx")

ITEM_RE = re.compile(r"^\s*-\s*\[(?: |x|X)\]\s+(\S.*)$")
HEADING_RE = re.compile(r"^###\s+(.+?)\s*$")
CODE_RE = re.compile(r"^(\d+(?:\.\d+)*)\s+(.*)$")
INVALID_SHEET_CHARS_RE = re.compile(r"[\[\]:*?/\\]")
URL_RE = re.compile(r"https?://[^\s>\)]+")


def sanitize_sheet_name(name: str, used: set[str]) -> str:
    base = INVALID_SHEET_CHARS_RE.sub("", name).strip() or "Sheet"
    base = base[:31]
    if base not in used:
        used.add(base)
        return base

    i = 2
    while True:
        suffix = f"_{i}"
        candidate = (base[: 31 - len(suffix)] + suffix).strip()
        if candidate not in used:
            used.add(candidate)
            return candidate
        i += 1


def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def clean_line_for_text(line: str) -> str:
    return re.sub(r"^\s*>\s?", "", line).strip()


def parse_markdown(md_text: str):
    materias: OrderedDict[str, list[dict]] = OrderedDict()
    retificacoes: list[dict] = []

    current_materia: str | None = None
    current_item_parts: list[str] | None = None

    lines = md_text.splitlines()
    i = 0

    def flush_item() -> None:
        nonlocal current_item_parts
        if current_materia is None or not current_item_parts:
            current_item_parts = None
            return

        text = normalize_spaces(" ".join(current_item_parts))
        if not text:
            current_item_parts = None
            return

        m = CODE_RE.match(text)
        if m:
            codigo = m.group(1)
            conteudo = m.group(2).strip()
            nivel = codigo.count(".") + 1
        else:
            codigo = ""
            conteudo = text
            nivel = ""

        materias.setdefault(current_materia, []).append(
            {
                "Status": "☐",
                "Codigo": codigo,
                "Conteudo": conteudo,
                "Nivel": nivel,
            }
        )
        current_item_parts = None

    while i < len(lines):
        line = lines[i]
        line_lc = line.lower()

        h = HEADING_RE.match(line)
        if h:
            flush_item()
            current_materia = h.group(1).strip()
            materias.setdefault(current_materia, [])
            i += 1
            continue

        if "retificado" in line_lc:
            flush_item()
            block = [clean_line_for_text(line)]
            i += 1
            while i < len(lines):
                nxt = lines[i]
                if HEADING_RE.match(nxt) or ITEM_RE.match(nxt):
                    break
                if not nxt.strip() and i + 1 < len(lines):
                    lookahead = lines[i + 1]
                    if HEADING_RE.match(lookahead) or ITEM_RE.match(lookahead):
                        break
                cleaned = clean_line_for_text(nxt)
                if cleaned:
                    block.append(cleaned)
                i += 1

            quote_text = normalize_spaces(" ".join(block))
            urls = URL_RE.findall(quote_text)
            retificacoes.append(
                {
                    "Materia": current_materia or "",
                    "Texto_Retificacao": quote_text,
                    "Fonte": " | ".join(urls),
                }
            )
            continue

        m_item = ITEM_RE.match(line)
        if m_item:
            flush_item()
            current_item_parts = [m_item.group(1).strip()]
            i += 1
            continue

        if current_item_parts is not None:
            stripped = line.strip()
            if (
                stripped
                and not HEADING_RE.match(line)
                and not ITEM_RE.match(line)
                and "retificado" not in line_lc
            ):
                current_item_parts.append(stripped)

        i += 1

    flush_item()
    return materias, retificacoes


def apply_table_format(sheet, max_col: int) -> None:
    sheet.freeze_panes = "A2"
    end_col = ["A", "B", "C", "D"][max_col - 1]
    sheet.auto_filter.ref = f"A1:{end_col}{sheet.max_row}"
    for cell in sheet[1]:
        cell.font = Font(bold=True)


def set_column_widths(sheet) -> None:
    widths = {"A": 10, "B": 12, "C": 120, "D": 8}
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width


def main() -> None:
    md_text = INPUT_MD.read_text(encoding="utf-8")
    materias, retificacoes = parse_markdown(md_text)

    wb = Workbook()
    wb.remove(wb.active)

    used_sheet_names: set[str] = set()

    for materia, itens in materias.items():
        title = sanitize_sheet_name(materia, used_sheet_names)
        ws = wb.create_sheet(title=title)
        ws.append(["Status", "Codigo", "Conteudo", "Nivel"])
        for item in itens:
            ws.append([item["Status"], item["Codigo"], item["Conteudo"], item["Nivel"]])
        apply_table_format(ws, 4)
        set_column_widths(ws)

    ws_ret = wb.create_sheet(title=sanitize_sheet_name("RETIFICACOES", used_sheet_names))
    ws_ret.append(["Materia", "Texto_Retificacao", "Fonte"])
    for r in retificacoes:
        ws_ret.append([r["Materia"], r["Texto_Retificacao"], r["Fonte"]])
    ws_ret.freeze_panes = "A2"
    ws_ret.auto_filter.ref = f"A1:C{ws_ret.max_row}"
    for cell in ws_ret[1]:
        cell.font = Font(bold=True)
    ws_ret.column_dimensions["A"].width = 40
    ws_ret.column_dimensions["B"].width = 140
    ws_ret.column_dimensions["C"].width = 60

    wb.save(OUTPUT_XLSX)

    print(f"Materias: {len(materias)}")
    print(f"Retificacoes: {len(retificacoes)}")
    print(f"Arquivo: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
